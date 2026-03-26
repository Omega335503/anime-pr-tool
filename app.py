"""アニメ広報レポート自動生成ツール（SSEストリーミング対応・ローカル開発用）"""

import os
import time
import re
import json
from datetime import datetime
from urllib.parse import unquote
from flask import Flask, render_template, request, Response, stream_with_context
import requests as http_requests
from bs4 import BeautifulSoup
from google import genai

app = Flask(__name__)

# Geminiクライアントをシングルトンで使い回す
_gemini_client = None
def get_gemini_client():
    global _gemini_client
    # GEMINI_API_KEYを明示的に使う（GOOGLE_API_KEYはSheets用）
    api_key = os.environ.get("GEMINI_API_KEY", "")
    if not api_key:
        api_key = os.environ.get("GOOGLE_API_KEY", "")
    if not api_key:
        return None
    if _gemini_client is None:
        _gemini_client = genai.Client(api_key=api_key)
    return _gemini_client

# --- generate.py と同じロジック（ローカル開発用に同期） ---

MEDIA_MAP = {
    "natalie.mu": "コミックナタリー",
    "comic-natalie.com": "コミックナタリー",
    "dengekionline.com": "電撃オンライン",
    "mantan-web.jp": "MANTANWEB",
    "anime.eiga.com": "アニメハック",
    "news.yahoo.co.jp": "Yahoo!ニュース",
    "animatetimes.com": "アニメイトタイムズ",
    "famitsu.com": "ファミ通.com",
    "4gamer.net": "4Gamer.net",
    "gigazine.net": "GIGAZINE",
    "nlab.itmedia.co.jp": "ねとらぼ",
    "oricon.co.jp": "ORICON NEWS",
    "realsound.jp": "リアルサウンド",
    "animeanime.jp": "アニメ!アニメ!",
    "eiga.com": "映画.com",
    "cinematoday.jp": "シネマトゥデイ",
    "jp.ign.com": "IGN Japan",
    "kai-you.net": "KAI-YOU",
    "game.watch.impress.co.jp": "GAME Watch",
    "hobby.watch.impress.co.jp": "HOBBY Watch",
    "av.watch.impress.co.jp": "AV Watch",
    "itmedia.co.jp": "ITmedia",
    "news.mynavi.jp": "マイナビニュース",
    "gamer.ne.jp": "Gamer",
    "comic-walker.com": "カドコミ",
    "koubo.jp": "Koubo",
    "news.toremaga.com": "とれまがニュース",
    "news.livedoor.com": "ライブドアニュース",
    "news.nifty.com": "ニフティニュース",
    "mdpr.jp": "モデルプレス",
    "excite.co.jp": "エキサイトニュース",
    "anime-recorder.com": "アニメレコーダー",
    "webnewtype.com": "WebNewtype",
    "febri.jp": "Febri",
    "lisani.jp": "リスアニ!",
    "nijimen.net": "にじめん",
    "hobby.dengeki.com": "電撃ホビーウェブ",
    "akiba-souken.com": "アキバ総研",
    "s-manga.net": "集英社マンガ",
    "websunday.net": "WEBサンデー",
    "bs4.jp": "BS日テレ",
    "ntv.co.jp": "日本テレビ",
}

PR_DOMAINS = {
    "prtimes.jp": "PR TIMES",
    "atpress.ne.jp": "@Press",
    "valuepress.com": "ValuePress!",
    "dreamnews.jp": "DreamNews",
}

SNS_DOMAINS = {"x.com", "twitter.com"}
INFO_DOMAINS = {"dic.pixiv.net", "ja.wikipedia.org", "anidb.net", "myanimelist.net",
                "anilist.co", "anime-planet.com", "annict.com"}
OFFICIAL_DOMAINS = {"anime.nicovideo.jp", "abema.tv", "tver.jp",
                    "netflix.com", "amazon.co.jp", "crunchyroll.com"}

SEARCH_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/131.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ja,en;q=0.9",
    "Accept-Encoding": "gzip, deflate",
}


def log(msg):
    print(f"[generate] {msg}", flush=True)


def get_domain(url):
    match = re.search(r'https?://(?:www\.)?([^/]+)', url)
    return match.group(1) if match else ""


def search_startpage(query, max_results=12, retries=1):
    for attempt in range(retries + 1):
        try:
            r = http_requests.post(
                "https://www.startpage.com/sp/search",
                data={"query": query, "cat": "web", "language": "japanese"},
                headers=SEARCH_HEADERS,
                timeout=15,
            )
            log(f"Startpage [{query[:30]}...] status={r.status_code} len={len(r.text)}")
            if r.status_code != 200:
                raise Exception(f"Status {r.status_code}")

            soup = BeautifulSoup(r.text, "html.parser")
            results = []
            for el in soup.select(".result")[:max_results]:
                upper = el.select_one(".upper a[href^='http']")
                if not upper:
                    upper = el.select_one("a[href^='http']")
                if not upper:
                    continue
                url = upper.get("href", "")
                if "startpage.com" in url:
                    continue

                title_el = el.select_one("a.result-title")
                title = title_el.get_text(strip=True) if title_el else ""

                snippet_el = el.select_one("p.description") or el.select_one("p")
                snippet = snippet_el.get_text(strip=True) if snippet_el else ""
                snippet = re.sub(r'^\d+\s*(?:日|時間|分|秒|週間|ヶ月)\s*前\.{3}\s*', '', snippet)

                if url and (title or snippet):
                    results.append({"href": url, "title": title, "body": snippet})

            log(f"  → {len(results)} results parsed")
            return results
        except Exception as e:
            log(f"Startpage attempt {attempt+1} failed: {e}")
            if attempt < retries:
                time.sleep(2)
    return []


def search_duckduckgo(query, retries=1):
    for attempt in range(retries + 1):
        try:
            r = http_requests.post(
                "https://html.duckduckgo.com/html/",
                data={"q": query, "kl": "jp-jp"},
                headers=SEARCH_HEADERS,
                timeout=15,
            )
            log(f"DDG [{query[:30]}...] status={r.status_code}")
            if r.status_code != 200:
                raise Exception(f"Status {r.status_code}")
            soup = BeautifulSoup(r.text, "html.parser")
            if "captcha" in r.text.lower() or "bots use" in r.text.lower():
                log("  DDG CAPTCHA detected")
                return []
            results = []
            titles = soup.select(".result__a")
            snippets = soup.select(".result__snippet")
            for i, a_tag in enumerate(titles):
                href = a_tag.get("href", "")
                if "uddg=" in href:
                    match = re.search(r'uddg=([^&]+)', href)
                    if match:
                        href = unquote(match.group(1))
                title = a_tag.get_text(strip=True)
                body = snippets[i].get_text(strip=True) if i < len(snippets) else ""
                if href and title:
                    results.append({"href": href, "title": title, "body": body})
            log(f"  → {len(results)} results")
            return results
        except Exception as e:
            log(f"DDG attempt {attempt+1} failed: {e}")
            if attempt < retries:
                time.sleep(2)
    return []


def search_web(query, retries=1):
    results = search_startpage(query, retries=retries)
    if results:
        return results
    log("Startpage failed, trying DuckDuckGo fallback...")
    return search_duckduckgo(query, retries=retries)


def classify_result(url, title, body, anime_title):
    domain = get_domain(url)
    ws_re = re.compile(r'[\s\u3000\u00a0\u200b]+')
    clean_anime = ws_re.sub('', anime_title)
    clean_title = ws_re.sub('', title)
    clean_body = ws_re.sub('', body)
    clean_url = ws_re.sub('', unquote(url))

    if clean_anime not in clean_title and clean_anime not in clean_body and clean_anime not in clean_url:
        return None, None

    for pr_domain, pr_name in PR_DOMAINS.items():
        if pr_domain in domain:
            return "press_release", {"source": pr_name, "title": title, "url": url, "body": body}

    if any(sns in domain for sns in SNS_DOMAINS):
        is_post = "/status/" in url
        account_match = re.search(r'(?:x\.com|twitter\.com)/(\w+)', url)
        account = f"@{account_match.group(1)}" if account_match else ""
        return "sns", {"account": account, "title": title, "url": url, "is_post": is_post, "body": body}

    if any(info in domain for info in INFO_DOMAINS):
        return "info", {"source": title, "url": url, "body": body}

    for media_domain, media_name in MEDIA_MAP.items():
        if media_domain in domain:
            return "media", {"media": media_name, "title": title, "url": url, "body": body}

    if any(kw in url for kw in ["/news/", "/article/", "/press/", "/topics/"]):
        return "media", {"media": domain, "title": title, "url": url, "body": body}

    if any(d in domain for d in OFFICIAL_DOMAINS):
        return "info", {"source": title, "url": url, "body": body}

    return "info", {"source": title, "url": url, "body": body}


def generate_report(anime_title, press_releases, media_coverage, sns_posts, info_pages):
    now = datetime.now().strftime("%Y年%m月%d日")
    lines = []
    lines.append(f"# 『{anime_title}』 広報レポート\n")
    lines.append(f"**作成日:** {now}")
    lines.append(f"**対象作品:** {anime_title}")
    lines.append("")
    lines.append("---\n")

    lines.append("## 1. プレスリリース\n")
    if press_releases:
        lines.append("| # | 配信元 | タイトル | URL |")
        lines.append("|---|--------|---------|-----|")
        for i, pr in enumerate(press_releases, 1):
            t = pr["title"][:60] + "..." if len(pr["title"]) > 60 else pr["title"]
            lines.append(f'| {i} | {pr["source"]} | {t} | {pr["url"]} |')
    else:
        lines.append("該当するプレスリリースは見つかりませんでした。\n")
    lines.append("")

    lines.append("---\n")
    lines.append("## 2. メディア掲載一覧\n")
    if media_coverage:
        lines.append("| # | メディア名 | 記事タイトル | URL |")
        lines.append("|---|----------|------------|-----|")
        for i, mc in enumerate(media_coverage, 1):
            t = mc["title"][:50] + "..." if len(mc["title"]) > 50 else mc["title"]
            lines.append(f'| {i} | {mc["media"]} | {t} | {mc["url"]} |')
        lines.append("")
        lines.append(f"**掲載メディア数: {len(set(mc['media'] for mc in media_coverage))}媒体 / 記事数: {len(media_coverage)}件**")
    else:
        lines.append("該当するメディア掲載は見つかりませんでした。\n")
    lines.append("")

    lines.append("---\n")
    lines.append("## 3. SNS（X / Twitter）\n")
    posts = [s for s in sns_posts if s["is_post"]]
    profiles = [s for s in sns_posts if not s["is_post"]]

    if posts:
        lines.append("### 関連投稿\n")
        lines.append("| # | アカウント | 内容 | URL |")
        lines.append("|---|----------|------|-----|")
        for i, p in enumerate(posts, 1):
            b = p["body"][:40] + "..." if len(p["body"]) > 40 else p["body"]
            lines.append(f'| {i} | {p["account"]} | {b} | {p["url"]} |')
    else:
        lines.append("関連投稿は検索で見つかりませんでした。\n")

    if profiles:
        lines.append("\n### 関連アカウント\n")
        for p in profiles:
            lines.append(f'- {p["account"]}: {p["url"]}')
    lines.append("")

    if info_pages:
        lines.append("---\n")
        lines.append("## 4. 関連情報ページ\n")
        for ip in info_pages:
            lines.append(f'- [{ip["source"]}]({ip["url"]})')
        lines.append("")

    lines.append("---\n")
    lines.append("## サマリー\n")
    lines.append(f"- プレスリリース: **{len(press_releases)}件**")
    mc_count = len(set(mc["media"] for mc in media_coverage)) if media_coverage else 0
    lines.append(f"- メディア掲載: **{mc_count}媒体 / {len(media_coverage)}記事**")
    lines.append(f"- SNS投稿: **{len(posts)}件**")
    lines.append(f"- 関連情報ページ: **{len(info_pages)}件**")
    lines.append("")

    lines.append("---\n")
    lines.append("## 注記\n")
    lines.append("- X（Twitter）のエンゲージメント数値（いいね・RT・インプレッション）はWeb検索では取得できません。")
    lines.append("- PR TIMESのPV数は管理画面からの確認が必要です。")
    lines.append("- 検索結果はStartpage (Google) のインデックスに依存しています。")
    lines.append("")
    lines.append("---\n")
    lines.append(f"*本レポートは自動検索により {now} に生成されました。*")

    return "\n".join(lines)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/generate", methods=["POST"])
def generate():
    auth_error = require_auth(request)
    if auth_error:
        return json.dumps(auth_error[0], ensure_ascii=False), auth_error[1]

    data = request.get_json()
    anime_title = data.get("title", "").strip()

    if not anime_title:
        return json.dumps({"error": "アニメタイトルを入力してください"}), 400

    log(f"=== START: '{anime_title}' ===")

    def stream():
        press_releases, media_coverage, sns_posts, info_pages = [], [], [], []
        seen_urls = set()

        steps = [
            ("アニメ関連記事を検索中...",
             f'"{anime_title}" アニメ ニュース'),
            ("メディア掲載を検索中...",
             f'{anime_title} site:natalie.mu OR site:animeanime.jp OR site:oricon.co.jp OR site:mantan-web.jp OR site:animatetimes.com'),
            ("プレスリリースを検索中...",
             f'"{anime_title}" site:prtimes.jp OR site:atpress.ne.jp'),
            ("追加ニュースを検索中...",
             f'"{anime_title}" プレスリリース OR 発表 OR 放送'),
            ("SNS投稿を検索中...",
             f'{anime_title} site:x.com'),
        ]

        total_steps = len(steps) + 1

        for i, (label, query) in enumerate(steps):
            yield f"data: {json.dumps({'type': 'progress', 'step': i + 1, 'total': total_steps, 'message': label}, ensure_ascii=False)}\n\n"

            if i > 0:
                time.sleep(2)

            results = search_web(query)

            for r in results:
                url = r.get("href", "")
                if url in seen_urls:
                    continue
                seen_urls.add(url)
                category, item = classify_result(url, r.get("title", ""), r.get("body", ""), anime_title)
                if category == "press_release":
                    press_releases.append(item)
                elif category == "media":
                    media_coverage.append(item)
                elif category == "sns":
                    sns_posts.append(item)
                elif category == "info":
                    info_pages.append(item)

        yield f"data: {json.dumps({'type': 'progress', 'step': total_steps, 'total': total_steps, 'message': 'レポートを生成中...'}, ensure_ascii=False)}\n\n"

        report = generate_report(anime_title, press_releases, media_coverage, sns_posts, info_pages)

        yield f"data: {json.dumps({'type': 'done', 'report': report, 'stats': {'press_releases': len(press_releases), 'media_coverage': len(media_coverage), 'sns_posts': len([s for s in sns_posts if s['is_post']]), 'info_pages': len(info_pages)}}, ensure_ascii=False)}\n\n"

        log(f"=== DONE: '{anime_title}' PR={len(press_releases)} Media={len(media_coverage)} SNS={len(sns_posts)} Info={len(info_pages)} ===")

    return Response(stream_with_context(stream()), content_type="text/event-stream")


# ===== コンテンツ生成（content_generate.py と同じロジック） =====

from api.content_generate import build_prompt
from api.fetch_gdoc import fetch_google_doc
from api.extract_info import build_individual_prompt, build_meta_prompt, extract_json_from_response, DOC_LABELS
from api.auth import require_auth


@app.route("/api/content_generate", methods=["POST"])
def content_generate():
    auth_error = require_auth(request)
    if auth_error:
        return json.dumps(auth_error[0], ensure_ascii=False), auth_error[1]

    data = request.get_json()
    content_type = data.get("content_type", "").strip()
    form_data = data.get("form_data", {})

    if content_type not in ("tweet", "press_release"):
        return json.dumps({"error": "content_type は 'tweet' または 'press_release' を指定してください"}), 400

    if not form_data.get("title"):
        return json.dumps({"error": "作品タイトルを入力してください"}), 400

    api_key = os.environ.get("GEMINI_API_KEY", "")
    if not api_key:
        return json.dumps({"error": "GEMINI_API_KEY が設定されていません。環境変数を設定してください。"}), 500

    log(f"=== CONTENT START: type={content_type} title='{form_data.get('title')}' ===")

    try:
        prompt = build_prompt(content_type, form_data)
    except ValueError as e:
        return json.dumps({"error": str(e)}), 400

    def stream():
        yield f"data: {json.dumps({'type': 'progress', 'message': 'Gemini APIに接続中...'}, ensure_ascii=False)}\n\n"

        try:
            client = get_gemini_client()

            full_text = ""
            response = client.models.generate_content_stream(
                model="gemini-2.5-flash",
                contents=prompt,
            )
            for chunk in response:
                if chunk.text:
                    full_text += chunk.text
                    yield f"data: {json.dumps({'type': 'chunk', 'text': chunk.text}, ensure_ascii=False)}\n\n"

            yield f"data: {json.dumps({'type': 'done', 'full_text': full_text}, ensure_ascii=False)}\n\n"
            log(f"=== CONTENT DONE: {len(full_text)} chars ===")

        except Exception as e:
            log(f"Error: {e}")
            yield f"data: {json.dumps({'type': 'error', 'message': f'Gemini API エラー: {str(e)}'}, ensure_ascii=False)}\n\n"

    return Response(stream_with_context(stream()), content_type="text/event-stream")


# ===== Google Docs/Sheets取得 =====

@app.route("/api/fetch_gdoc", methods=["POST"])
def api_fetch_gdoc():
    t0 = time.time()
    auth_error = require_auth(request)
    if auth_error:
        return json.dumps(auth_error[0], ensure_ascii=False), auth_error[1]
    log(f"[TIMING] fetch_gdoc auth: {time.time()-t0:.2f}s")

    data = request.get_json()
    url = data.get("url", "").strip()
    if not url:
        return json.dumps({"error": "URLを指定してください"}), 400
    try:
        t1 = time.time()
        result = fetch_google_doc(url)
        log(f"[TIMING] fetch_gdoc fetch: {time.time()-t1:.2f}s, content_len={len(result.get('content',''))}")
        log(f"[TIMING] fetch_gdoc TOTAL: {time.time()-t0:.2f}s")
        return json.dumps(result, ensure_ascii=False)
    except ValueError as e:
        return json.dumps({"error": str(e)}), 400
    except Exception as e:
        log(f"fetch_gdoc error: {e}")
        return json.dumps({"error": f"取得エラー: {str(e)}"}), 500


# ===== 情報抽出 =====

@app.route("/api/extract_info", methods=["POST"])
def api_extract_info():
    auth_error = require_auth(request)
    if auth_error:
        return json.dumps(auth_error[0], ensure_ascii=False), auth_error[1]

    data = request.get_json()
    documents = data.get("documents", {})

    if not any(documents.get(k, {}).get("content") for k in ("timeline", "text_master", "budget")):
        return json.dumps({"error": "少なくとも1つの資料が必要です"}), 400

    api_key = os.environ.get("GEMINI_API_KEY", "")
    if not api_key:
        return json.dumps({"error": "GEMINI_API_KEY が設定されていません"}), 500

    log("=== EXTRACT START ===")
    try:
        client = get_gemini_client()
        structured = {}

        # 各資料を個別に構造化
        for doc_type in ("timeline", "text_master", "budget"):
            doc = documents.get(doc_type)
            if not doc or not doc.get("content"):
                continue

            prompt = build_individual_prompt(doc_type, doc["content"])
            if not prompt:
                continue

            log(f"Structuring {doc_type}... ({len(doc['content'])} chars)")
            response = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=prompt,
            )
            structured[doc_type] = response.text.strip()
            log(f"  → {len(structured[doc_type])} chars")

        # 基本メタ情報を抽出
        log("Extracting meta...")
        meta_prompt = build_meta_prompt(documents)
        meta_response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=meta_prompt,
        )
        try:
            meta = extract_json_from_response(meta_response.text)
        except (json.JSONDecodeError, Exception):
            meta = {"title": "", "hashtags": "", "official_url": "", "official_x": ""}

        log(f"=== EXTRACT DONE: title={meta.get('title', '?')} ===")

        return json.dumps({"extracted": meta, "structured": structured}, ensure_ascii=False)

    except Exception as e:
        log(f"extract_info error: {e}")
        return json.dumps({"error": f"抽出エラー: {str(e)}"}), 500


# ===== 個別資料構造化 =====

def parse_csv_to_table(content):
    """CSVテキストを直接パースしてテーブルJSON構造を生成（AI不要・高速）"""
    import csv
    import io

    reader = csv.reader(io.StringIO(content))
    all_rows = list(reader)
    if not all_rows:
        return {"headers": [], "release_columns": [], "sections": []}

    # ヘッダー行（1行目）
    raw_headers = all_rows[0]

    # 不要な列を検出して除外
    skip_cols = set()
    max_col = max(len(r) for r in all_rows) if all_rows else 0
    for ci in range(max_col):
        h_strip = raw_headers[ci].strip() if ci < len(raw_headers) else ''
        # ヘッダーが空 or 「：」「:」
        if h_strip in ('：', ':', ''):
            col_vals = [r[ci].strip() if ci < len(r) else '' for r in all_rows[1:] if r]
            non_empty = [v for v in col_vals if v and v not in ('：', ':', '')]
            # 列の中身がほぼ「：」か空なら除外
            if len(non_empty) < max(1, len(col_vals) * 0.2):
                skip_cols.add(ci)

    # フィルタリング後のヘッダー
    headers = [h for ci, h in enumerate(raw_headers) if ci not in skip_cols]

    # 解禁列を検出（ヘッダーに日付・施策名っぽいもの）
    release_columns = []
    for ci, h in enumerate(headers):
        h_strip = h.strip()
        if h_strip and h_strip != headers[0] and (
            '月' in h_strip or '/' in h_strip or
            h_strip.replace('.', '').isdigit() or
            any(kw in h_strip for kw in ['PV', 'HP', '話', '放送', 'CM'])
        ):
            release_columns.append(ci)

    # データ行をセクション分割
    sections = []
    current_section = {"title": "", "rows": []}

    for row_idx, raw_row in enumerate(all_rows[1:]):
        # skip_cols除外
        row = [c for ci, c in enumerate(raw_row) if ci not in skip_cols]
        # 列数をheadersに揃える
        while len(row) < len(headers):
            row.append('')

        # 空行チェック → 次にセクションヘッダーが来るかもしれないのでフラグだけ立てる
        non_empty_cells = [c.strip() for c in row if c.strip()]
        if not non_empty_cells:
            continue

        # セクションヘッダー検出
        # 条件: 1列目だけに短いテキストがあり、他が全部空で、●もない
        first_val = row[0].strip()
        other_vals = [c.strip() for c in row[1:] if c.strip()]
        is_header_candidate = (
            first_val
            and not other_vals
            and len(first_val) < 20
            and first_val not in ('●', '⚫︎')
            and not any(c in first_val for c in ('：', ':', '/', 'ⓒ', '©'))
            and not any(first_val.startswith(kw) for kw in ('20', 'http'))
        )
        if is_header_candidate:
            if current_section["rows"]:
                sections.append(current_section)
            current_section = {"title": first_val, "rows": []}
            continue

        # 通常データ行
        current_section["rows"].append(row[:len(headers)])

    # 最後のセクション
    if current_section["rows"]:
        sections.append(current_section)

    # titleがない最初のセクションにデフォルト名
    if sections and not sections[0]["title"]:
        sections[0]["title"] = "作品情報"

    # 全セクションの行をキーワードで再分類
    ROW_KEYWORDS = {
        "楽曲": ["主題歌", "オープニング", "エンディング", "OP曲", "ED曲", "挿入歌"],
        "SNS": ["HP", " X ", "Twitter", "Instagram", "YouTube", "TikTok"],
        "制作物": ["ビジュアル", "KV", "PV", "CM", "番宣", "キービジュアル"],
        "権利表記": ["マルシー", "ⓒ", "©"],
    }

    # まず全行をフラットに、元のセクション情報付きで集める
    all_categorized = {}  # {category: [rows]}
    for section in sections:
        for row in section["rows"]:
            labels_text = row[0].strip() + ' ' + row[1].strip() if len(row) > 1 else row[0].strip()

            # キーワードマッチ
            matched_cat = None
            for cat_name, keywords in ROW_KEYWORDS.items():
                if any(kw in labels_text for kw in keywords):
                    matched_cat = cat_name
                    break

            # 放送情報（年号+放送）→ 作品情報に統合
            if not matched_cat and ('放送' in labels_text or '配信' in labels_text):
                matched_cat = "作品情報"

            # マルシー等を作品情報に
            if matched_cat == "権利表記":
                matched_cat = "作品情報"

            # キーワードマッチ優先、なければ元のセクション名
            if matched_cat:
                target = matched_cat
            elif section["title"]:
                # 元セクション名もキーワードで再分類
                sec_title = section["title"]
                if sec_title in ("X", "HP", "SNS", "Twitter"):
                    target = "SNS"
                elif sec_title in ("PV1", "PV2", "PV", "CM"):
                    target = "制作物"
                else:
                    target = sec_title
            else:
                target = "その他"

            if target not in all_categorized:
                all_categorized[target] = []
            all_categorized[target].append(row)

    # セクション順序を決定
    SECTION_ORDER = ["作品情報", "スタッフ", "キャスト", "楽曲", "SNS", "制作物", "その他"]
    sections = []
    seen = set()
    for name in SECTION_ORDER:
        if name in all_categorized:
            sections.append({"title": name, "rows": all_categorized[name]})
            seen.add(name)
    # 残り
    for name, rows in all_categorized.items():
        if name not in seen:
            sections.append({"title": name, "rows": rows})

    # 解禁列の右側自動連鎖: ●がある列より右側は全て●にする
    if release_columns:
        sorted_rel = sorted(release_columns)
        for section in sections:
            for row in section["rows"]:
                first_release_idx = None
                for rc in sorted_rel:
                    if rc < len(row) and row[rc] == '●':
                        first_release_idx = rc
                        break
                if first_release_idx is not None:
                    for rc in sorted_rel:
                        if rc >= first_release_idx:
                            if rc < len(row):
                                row[rc] = '●'

    return {
        "headers": headers,
        "release_columns": release_columns,
        "sections": sections,
    }


@app.route("/api/structure_doc", methods=["POST"])
def api_structure_doc():
    t0 = time.time()
    auth_error = require_auth(request)
    if auth_error:
        return json.dumps(auth_error[0], ensure_ascii=False), auth_error[1]

    data = request.get_json()
    doc_type = data.get("doc_type", "")
    content = data.get("content", "")
    instructions = data.get("instructions", "")

    if doc_type not in ("timeline", "text_master", "budget"):
        return json.dumps({"error": "無効なdoc_typeです"}), 400
    if not content:
        return json.dumps({"error": "contentが空です"}), 400

    try:
        if doc_type in ("text_master", "budget"):
            t1 = time.time()
            parsed = parse_csv_to_table(content)
            log(f"[TIMING] structure_doc {doc_type} direct_parse: {time.time()-t1:.3f}s")
            log(f"[TIMING] structure_doc {doc_type} TOTAL: {time.time()-t0:.3f}s")
            return json.dumps({"structured": parsed}, ensure_ascii=False)
        else:
            api_key = os.environ.get("GEMINI_API_KEY", "")
            if not api_key:
                return json.dumps({"error": "GEMINI_API_KEY が設定されていません"}), 500

            prompt = build_individual_prompt(doc_type, content)
            if instructions:
                prompt += f"\n\n## ユーザーからの追加指示（必ず従うこと）\n{instructions}"
                log(f"[structure_doc] Extra instructions: {instructions}")
            log(f"[TIMING] structure_doc {doc_type} prompt_len={len(prompt)}")

            t2 = time.time()
            client = get_gemini_client()
            response = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=prompt,
            )
            log(f"[TIMING] structure_doc {doc_type} gemini: {time.time()-t2:.2f}s")
            log(f"[TIMING] structure_doc {doc_type} TOTAL: {time.time()-t0:.2f}s")

            try:
                parsed = extract_json_from_response(response.text)
                return json.dumps({"structured": parsed}, ensure_ascii=False)
            except (json.JSONDecodeError, Exception):
                return json.dumps({"structured": response.text.strip()}, ensure_ascii=False)

    except Exception as e:
        log(f"structure_doc error: {e}")
        return json.dumps({"error": f"構造化エラー: {str(e)}"}), 500


# ===== AI自動入力 =====

@app.route("/api/ai_fill", methods=["POST"])
def api_ai_fill():
    auth_error = require_auth(request)
    if auth_error:
        return json.dumps(auth_error[0], ensure_ascii=False), auth_error[1]

    data = request.get_json()
    prompt = data.get("prompt", "")
    if not prompt:
        return json.dumps({"error": "promptが空です"}), 400

    try:
        client = get_gemini_client()
        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
        )
        text = response.text.strip()
        parsed = extract_json_from_response(text)
        return json.dumps({"fields": parsed}, ensure_ascii=False)
    except Exception as e:
        log(f"ai_fill error: {e}")
        return json.dumps({"error": str(e)}), 500


# ===== メタ情報抽出 =====

@app.route("/api/extract_meta", methods=["POST"])
def api_extract_meta():
    t0 = time.time()
    auth_error = require_auth(request)
    if auth_error:
        return json.dumps(auth_error[0], ensure_ascii=False), auth_error[1]
    log(f"[TIMING] extract_meta auth: {time.time()-t0:.2f}s")

    data = request.get_json()
    documents = data.get("documents", {})

    api_key = os.environ.get("GEMINI_API_KEY", "")
    if not api_key:
        return json.dumps({"error": "GEMINI_API_KEY が設定されていません"}), 500

    try:
        t1 = time.time()
        prompt = build_meta_prompt(documents)
        log(f"[TIMING] extract_meta prompt_build: {time.time()-t1:.2f}s, prompt_len={len(prompt)}")

        t2 = time.time()
        client = get_gemini_client()
        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
        )
        log(f"[TIMING] extract_meta gemini_call: {time.time()-t2:.2f}s")
        log(f"[TIMING] extract_meta TOTAL: {time.time()-t0:.2f}s")

        meta = extract_json_from_response(response.text)
        return json.dumps({"extracted": meta}, ensure_ascii=False)
    except Exception as e:
        log(f"extract_meta error: {e}")
        return json.dumps({"extracted": {"title": "", "hashtags": "", "official_url": "", "official_x": ""}}, ensure_ascii=False)


# ===== ガントチャートxlsx生成 =====

@app.route("/api/generate_gantt", methods=["POST"])
def api_generate_gantt():
    """タイムラインデータからガントチャートxlsxを生成"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file

    auth_error = require_auth(request)
    if auth_error:
        return json.dumps(auth_error[0], ensure_ascii=False), auth_error[1]

    data = request.get_json()
    title = data.get("title", "作品名未設定")
    timeline = data.get("timeline", [])  # [{date, title, category, items:[{text, children}]}]

    if not timeline:
        return json.dumps({"error": "タイムラインデータがありません"}), 400

    # --- 月の範囲を算出 ---
    import re as re_mod
    from datetime import datetime as dt_cls

    # タイムラインの日付からstart/endを推定
    dates = []
    for entry in timeline:
        d = entry.get("date", "")
        m = re_mod.match(r'(\d{1,2})/(\d{1,2})', d)
        if m:
            month = int(m.group(1))
            dates.append(month)

    if not dates:
        # デフォルト: 現在月から18ヶ月
        now = dt_cls.now()
        start_month = now.month
        start_year = now.year
    else:
        start_month = min(dates)
        start_year = dt_cls.now().year

    # 24ヶ月分のカラムを生成
    TOTAL_MONTHS = 24
    months = []
    y, m = start_year, start_month
    for _ in range(TOTAL_MONTHS):
        months.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1

    # --- カテゴリ別にイベントを分類 ---
    CATEGORIES = ["プレスリリース", "WEB/HP", "イベント", "SNS更新", "制作物", "アニメ制作", "MTG", "その他"]
    CATEGORY_MAP = {
        "プレスリリース": "プレスリリース",
        "アニメ制作": "アニメ制作",
        "MTG": "MTG",
        "SNS更新": "SNS更新",
        "イベント": "イベント",
        "WEB/HP": "WEB/HP",
        "制作物": "制作物",
        # 旧名互換
        "制作進行": "アニメ制作",
        "X更新": "SNS更新",
        "Web/HP": "WEB/HP",
        "PV/映像": "制作物",
    }

    cat_events = {c: [] for c in CATEGORIES}
    for entry in timeline:
        d = entry.get("date", "")
        cat = entry.get("category", "その他")
        mapped = CATEGORY_MAP.get(cat, cat)
        if mapped not in cat_events:
            mapped = "その他"

        event_title = entry.get("title", "")
        items = entry.get("items", [])
        items_text = [it.get("text", "") for it in items]

        m_match = re_mod.match(r'(\d{1,2})/(\d{1,2})', d)
        month = int(m_match.group(1)) if m_match else None
        day = int(m_match.group(2)) if m_match else None
        week = min((day - 1) // 7, 3) if day else 0

        cat_events[mapped].append({
            "month": month,
            "week": week,
            "title": event_title,
            "items": items_text,
        })

    # --- Excelワークブック生成 ---
    wb = Workbook()
    ws = wb.active
    ws.title = "宣伝スケジュール"

    # スタイル定義
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    week_font = Font(size=9)
    cat_font = Font(bold=True, size=10)
    event_font = Font(size=9)

    month_fill = PatternFill("solid", fgColor="1F4E79")
    month_font = Font(bold=True, color="FFFFFF", size=10)
    week_fill = PatternFill("solid", fgColor="000000")
    week_font_white = Font(color="FFFFFF", size=9)

    cat_fills = {
        "情報解禁": PatternFill("solid", fgColor="F4CCCC"),
        "公式WEBサイト": PatternFill("solid", fgColor="D9EAD3"),
        "イベント": PatternFill("solid", fgColor="CFE2F3"),
        "SNS": PatternFill("solid", fgColor="FCE5CD"),
        "X": PatternFill("solid", fgColor="D9D2E9"),
        "その他": PatternFill("solid", fgColor="EFEFEF"),
    }

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # 列A: 実施項目（幅広め）
    ws.column_dimensions['A'].width = 16

    # Row 1: タイトル
    ws.cell(row=1, column=1, value=f'TVアニメ「{title}」宣伝スケジュール').font = title_font

    # Row 2: 空行

    # Row 3: 年ヘッダー
    col = 2
    prev_year = None
    for y, m in months:
        if y != prev_year:
            ws.cell(row=3, column=col, value=f"{y}年").font = Font(bold=True)
            prev_year = y
        col += 4  # 各月4週

    # Row 4: 月ヘッダー
    col = 2
    for y, m in months:
        cell = ws.cell(row=4, column=col, value=f"{m}月")
        cell.font = month_font
        cell.fill = month_fill
        cell.alignment = Alignment(horizontal='center')
        # 4列分マージ
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+3)
        col += 4

    # Row 5: 週ヘッダー (1W-4W)
    col = 2
    for _ in months:
        for w in range(1, 5):
            cell = ws.cell(row=5, column=col, value=f"{w}W")
            cell.font = week_font_white
            cell.fill = week_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col)].width = 4
            col += 1

    # Row 6: 狙い (空、後で手動追加可能)
    ws.cell(row=6, column=1, value="狙い").font = Font(bold=True)

    # --- カテゴリ行を配置 ---
    current_row = 7
    for cat_name in CATEGORIES:
        events = cat_events.get(cat_name, [])

        # カテゴリヘッダー行
        cat_cell = ws.cell(row=current_row, column=1, value=cat_name)
        cat_cell.font = cat_font
        fill = cat_fills.get(cat_name, PatternFill("solid", fgColor="EFEFEF"))

        # カテゴリ行全体に背景色
        for c in range(1, 2 + len(months) * 4):
            ws.cell(row=current_row, column=c).fill = fill

        current_row += 1

        # イベントを配置
        if not events:
            current_row += 1  # 空行
            continue

        # 月+週ごとにグループ化
        placed = {}  # (month, week) -> [event_texts]
        for ev in events:
            key = (ev["month"], ev["week"])
            text = f"・{ev['title']}"
            if ev["items"]:
                for item in ev["items"][:2]:  # 最大2項目
                    text += f"\n  - {item}"
            if key not in placed:
                placed[key] = []
            placed[key].append(text)

        # 最大行数を計算
        max_items_in_row = max((len(v) for v in placed.values()), default=0)
        rows_needed = max(max_items_in_row, 1)

        for offset in range(rows_needed):
            for (month, week), texts in placed.items():
                if offset >= len(texts):
                    continue
                # 月のインデックスを見つける
                for mi, (y, m) in enumerate(months):
                    if m == month:
                        col_idx = 2 + mi * 4 + week
                        cell = ws.cell(row=current_row + offset, column=col_idx, value=texts[offset])
                        cell.font = event_font
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        break

        current_row += rows_needed + 1  # 空行を挟む

    # ウィンドウ枠を固定（A列と月/週ヘッダー）
    ws.freeze_panes = "B6"

    # --- 出力 ---
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{title}_宣伝スケジュール.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    app.run(debug=True, port=5001)
